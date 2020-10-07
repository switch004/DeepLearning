# -*- coding: utf-8 -*
import os
import re
import numpy as np
import six
from PIL import Image
from itertools import chain
import glob
import time
import matplotlib.pylab as plt
from sklearn.datasets import fetch_mldata
import chainer
import chainer.links as L
import chainer.functions as F
from chainer import Chain, optimizers, Variable, serializers, iterators, training, datasets
from chainer.datasets import LabeledImageDataset, tuple_dataset
from chainer.training import extensions, triggers 
from chainer.serializers import npz
from chainer.backends import cuda
from chainer import Function
from chainer import Link, ChainList
from chainer import cuda
from shuffle import shuffle
from augmentation import augmentation
from resize import resize
from excel import change_excel
from edit import edit_excel
from scale_augmentation import scale_data_augmentation
from count_image import count_image
from accuracy import accuracy
import xlwt
import shutil
import openpyxl as xl
from natsort import natsorted
from decimal import *



def image2Train(pathsAndLabels, channels=3):

    allData = []
    count_train_label = np.zeros(G)

    #データの追加(画像ファイル，ラベル)
    for pathsAndLabel in pathsAndLabels:
        path = pathsAndLabel[0]
        label = pathsAndLabel[1]
        imagelist = glob.glob(path + '*')
        count_train_label[int(label)] = len(imagelist)
        for imgName in imagelist:
            allData.append([imgName, label])
    allData = np.random.permutation(allData)#シャッフル

    #チャンネルが1の時は，画像ファイルとラベルデータに追加
    if channels == 1:
        imageData = []
        labelData = []
        for pathAndLabel in allData:
            img = Image.open(pathsAndLabel[0])
            imgData = np.asarray([np.float32(img)/255.0])#画像の正規化
            imageData.append(imgData)
            labelData.append(np.int32(pathsAndLabel[1]))

        threshold = np.int32(len(imageData)/8*7)#データの何割を教師データとテストデータにするか
        train = (imageData[0:threshold], labelData[0:threshold])
        test = (imageData[threshold:], labeData[threshold:])
    #RGBの時の処理
    else:
        imageData = []
        labelData = []
        for pathAndLabel in allData:
            img = Image.open(pathAndLabel[0])

            #新たに追加
            img = np.asarray(img, dtype=np.float32)
            img = img[:, :, ::-1]
            img -= np.array([103.939, 116.779, 123.68], dtype=np.float32)
            #print(img)
            img = img/255.0

            imageData.append(img)
            labelData.append(np.int32(pathAndLabel[1]))

        dataset = {}
        dataset['train_img'] = np.array(imageData[0:N]).transpose(0,3,1,2)
        dataset['train_label'] = np.array(labelData[0:N])


    return (dataset['train_img'], dataset['train_label']), count_train_label


def image2Test(pathsAndLabels, channels=3):

    allData = []
    count_test_label = np.zeros(G)

    #データの追加(画像ファイル，ラベル)
    for pathsAndLabel in pathsAndLabels:
        path = pathsAndLabel[0]
        label = pathsAndLabel[1]
        imagelist = glob.glob(path + '*')
        count_test_label[int(label)] = len(imagelist)
        for imgName in imagelist:
            allData.append([imgName, label])
    allData = np.random.permutation(allData)#シャッフル

    #チャンネルが1の時は，画像ファイルとラベルデータに追加
    if channels == 1:
        imageData = []
        labelData = []
        for pathAndLabel in allData:
            img = Image.open(pathsAndLabel[0])
            imgData = np.asarray([np.float32(img)/255.0])#画像の正規化
            imageData.append(imgData)
            labelData.append(np.int32(pathsAndLabel[1]))

        threshold = np.int32(len(imageData)/8*7)#データの何割を教師データとテストデータにするか
        train = (imageData[0:threshold], labelData[0:threshold])
        test = (imageData[threshold:], labeData[threshold:])
    #RGBの時の処理
    else:
        imageData = []
        labelData = []
        for pathAndLabel in allData:
            img = Image.open(pathAndLabel[0])

            #新たに追加
            img = np.asarray(img, dtype=np.float32)
            img = img[:, :, ::-1]
            img -= np.array([103.939, 116.779, 123.68], dtype=np.float32)
            img = img/255.0
            imageData.append(img)
            labelData.append(np.int32(pathAndLabel[1]))

        dataset = {}

        dataset['test_img'] = np.array(imageData[0:N_test]).transpose(0,3,1,2)
        dataset['test_label'] = np.array(labelData[0:N_test])


    return (dataset['test_img'], dataset['test_label']), count_test_label

class CNN(Chain):
    def __init__(self):
        super(CNN, self).__init__(
            conv1 = L.Convolution2D(3, 40, 15), # 入力:3 , 出力:40。両方記述。filter 5
            conv2 = L.Convolution2D(40, 50, 15), # 入力:40 , 出力:50。両方記述。filter 5
            l1 = L.Linear(None, 500),  # ここの入力の計算は以下の参考HPを参照。
            l2 = L.Linear(500, 500),
            l3 = L.Linear(500, G, initialW=np.zeros((G, 500), dtype=np.float32))#ここの値は設定する出力の数によって変更
        )
    def forward(self, x, ratio = 0.1):
        h = F.max_pooling_2d(F.relu(self.conv1(x)), 2)
        h = F.max_pooling_2d(F.relu(self.conv2(h)), 2)
        h = F.dropout(F.relu(self.l1(h)), ratio)
        h = F.dropout(F.relu(self.l2(h)), ratio)
        h = self.l3(h)
        return h


class VGG(chainer.Chain):
    def __init__(self, class_labels, pretrained_model='VGG_ILSVRC_16_layers.npz'):
        super(VGG, self).__init__()
        with self.init_scope():
            self.base = BaseVGG()
            self.fc6 = L.Linear(None, 793, initialW=chainer.initializers.HeNormal())
            self.fc7 = L.Linear(None, 252, initialW=chainer.initializers.HeNormal())
            self.fc8 = L.Linear(None, class_labels, initialW=chainer.initializers.HeNormal())
        npz.load_npz(pretrained_model, self.base)

    def __call__(self, x, t):
        h = self.predict(x)
        loss = F.softmax_cross_entropy(h,t)
        chainer.report({'loss': loss, 'accuracy': F.accuracy(h,t)}, self)
        return loss

    def predict(self, x):
        h = self.base(x)
        h = F.relu(self.fc6(h))
        h = F.relu(self.fc7(h))
        h = self.fc8(h)
        return h

class BaseVGG(Chain):
    def __init__(self):
        super(BaseVGG, self).__init__()
        with self.init_scope():
            self.conv1_1 = L.Convolution2D(None, 64, 3, 1, 1)
            self.conv1_2 = L.Convolution2D(64, 64, 3, 1, 1)
            self.conv2_1 = L.Convolution2D(64, 128, 3, 1, 1)
            self.conv2_2 = L.Convolution2D(128, 128, 3, 1, 1)
            self.conv3_1 = L.Convolution2D(128, 256, 3, 1, 1)
            self.conv3_2 = L.Convolution2D(256, 256, 3, 1, 1)
            self.conv3_3 = L.Convolution2D(256, 256, 3, 1, 1)
            self.conv4_1 = L.Convolution2D(256, 512, 3, 1, 1)
            self.conv4_2 = L.Convolution2D(512, 512, 3, 1, 1)
            self.conv4_3 = L.Convolution2D(512, 512, 3, 1, 1)
            self.conv5_1 = L.Convolution2D(512, 512, 3, 1, 1)
            self.conv5_2 = L.Convolution2D(512, 512, 3, 1, 1)
            self.conv5_3 = L.Convolution2D(512, 512, 3, 1, 1)


    def __call__(self, x):
        h = F.relu(self.conv1_1(x))
        h = F.relu(self.conv1_2(h))
        h = F.max_pooling_2d(h, ksize=2, stride=2)

        h = F.relu(self.conv2_1(h))
        h = F.relu(self.conv2_2(h))
        h = F.max_pooling_2d(h, ksize=2, stride=2)

        h = F.relu(self.conv3_1(h))
        h = F.relu(self.conv3_2(h))
        h = F.relu(self.conv3_3(h))
        h = F.max_pooling_2d(h, ksize=2, stride=2)

        h = F.relu(self.conv4_1(h))
        h = F.relu(self.conv4_2(h))
        h = F.relu(self.conv4_3(h))
        h = F.max_pooling_2d(h, ksize=2, stride=2)

        h = F.relu(self.conv5_1(h))
        h = F.relu(self.conv5_2(h))
        h = F.relu(self.conv5_3(h))
        h = F.max_pooling_2d(h, ksize=2, stride=2)

        return h


if __name__ == '__main__':
    print('class_labels変更した?')
    #以下パラメータの設定
    k = 10 #試行回数
    n_epoch = 20 #更新回数
    batch_size = 125 #バッチサイズ
    batch_size_predict = 1
    cnt_trial = 1#試行回数ごとの写真フォルダ分けに使用
    h = 90#入力画像の高さ
    w = 78#入力画像の幅
    gpu = 0 #GPUを用いるなら0,CPUなら-1
    N_test = 793#テスト画像枚数(ここの数字は再確認)

    #以下扱うディレクトリのパス設定
    input_dir = 'C:\\Users\\user\\Desktop\\input'#入力教師用画像フォルダ
    input_testing_dir = 'C:\\Users\\user\\Desktop\\input_test'#入力評価用画像フォルダ
    main_dir = 'C:\\Users\\user\\Desktop\\main'#mainフォルダ
    result_dir = 'C:\\Users\\user\\Desktop\\result'#結果フォルダ

    #main_dirがなければ作成
    if not os.path.exists(main_dir):
        os.mkdir(main_dir)

    #result_dirがなければ作成
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)

    #誤判別時の画像出力フォルダ作成
    miss_judge_train_dir = os.path.join(result_dir, 'miss_judge_train')
    miss_judge_test_dir = os.path.join(result_dir, 'miss_judge_test')

    #誤班別時画像出力フォルダがなければ，作成
    if not os.path.exists(miss_judge_train_dir):
        os.mkdir(miss_judge_train_dir)
    if not os.path.exists(miss_judge_test_dir):
        os.mkdir(miss_judge_test_dir)
    for j in range(1, k+1):
        if not os.path.exists(os.path.join(miss_judge_train_dir, str(j)+'_trial')):
            os.mkdir(os.path.join(miss_judge_train_dir, str(j)+'_trial'))
        if not os.path.exists(os.path.join(miss_judge_test_dir, str(j)+'_trial')):
            os.mkdir(os.path.join(miss_judge_test_dir, str(j)+'_trial'))
        for l in range(1, n_epoch+1):
            if not os.path.exists(os.path.join(os.path.join(miss_judge_train_dir, str(j)+'_trial'), str(l))):
                os.mkdir(os.path.join(os.path.join(miss_judge_train_dir, str(j)+'_trial'), str(l)))
            if not os.path.exists(os.path.join(os.path.join(miss_judge_test_dir, str(j)+'_trial'), str(l))):
                os.mkdir(os.path.join(os.path.join(miss_judge_test_dir, str(j)+'_trial'), str(l)))

    #正解時の画像出力フォルダ作成
    correct_judge_train_dir = os.path.join(result_dir, 'correct_judge_train')
    correct_judge_test_dir = os.path.join(result_dir, 'correct_judge_test')

    #正解判別時画像出力フォルダがなければ，作成
    if not os.path.exists(correct_judge_train_dir):
        os.mkdir(correct_judge_train_dir)
    if not os.path.exists(correct_judge_test_dir):
        os.mkdir(correct_judge_test_dir)
    for j in range(1,k+1):
        if not os.path.exists(os.path.join(correct_judge_train_dir, str(j)+'_trial')):
            os.mkdir(os.path.join(correct_judge_train_dir, str(j)+'_trial'))
        if not os.path.exists(os.path.join(correct_judge_test_dir, str(j)+'_trial')):
            os.mkdir(os.path.join(correct_judge_test_dir, str(j)+'_trial'))
        for l in range(1, n_epoch+1):
            if not os.path.exists(os.path.join(os.path.join(correct_judge_train_dir, str(j)+'_trial'), str(l))):
                os.mkdir(os.path.join(os.path.join(correct_judge_train_dir, str(j)+'_trial'), str(l)))
            if not os.path.exists(os.path.join(os.path.join(correct_judge_test_dir, str(j)+'_trial'), str(l))):
                os.mkdir(os.path.join(os.path.join(correct_judge_test_dir, str(j)+'_trial'), str(l)))

    #以下訓練と評価
    for i in range(1, k+1):
        start_time = time.clock()
        book = xlwt.Workbook()
        print(str(i) + '_trial')
        #試行ごとのフォルダ作成
        current_dir = os.path.join(main_dir, str(i)+'_trial')
        if not os.path.exists(current_dir):
            os.mkdir(current_dir)

        #試行ごとの結果フォルダ作成
        current_result_dir = os.path.join(result_dir, str(i)+'_trial')
        if not os.path.exists(current_result_dir):
            os.mkdir(current_result_dir)


        train_dir = os.path.join(current_dir, 'training')
        test_dir = os.path.join(current_dir, 'testing')

        #shuffleにより評価用画像枚数を返す．
        #N_testはテスト枚数,flagが0ならinputdirから教師用とテスト用に8:2でわける．flagが1なら教師用フォルダから教師用に，評価用フォルダから評価用にわける．

        #N_test = shuffle(input_dir, current_dir, flag = 0)
        N_test = shuffle(input_dir, current_dir, input_testing_dir, flag = 1)#N_test:評価用画像枚数
        
        #scale_augmentationの実行
        scale_data_augmentation(os.path.join(train_dir, 'image1'), os.path.join(train_dir, 'image2'), (h,w))



        #resizeの実行
        resize(os.path.join(train_dir, 'image2'), os.path.join(train_dir, 'image3'), h, w)
        resize(os.path.join(test_dir, 'image1'), os.path.join(test_dir, 'image2'), h, w)



        #data augmentationの実行
        augmentation(os.path.join(train_dir, 'image3'), os.path.join(train_dir, 'image4'), (h,w))

        #count_imageの実行
        #これにより教師用画像枚数とグループ数取得
        N, G = count_image(os.path.join(train_dir, 'image4'))
        print("N:" + str(N))
        print("N_test:" + str(N_test))

        #画像フォルダのパス
        IMG_DIR_train = os.path.join(train_dir, 'image4')
        IMG_DIR_test = os.path.join(test_dir, 'image2')


        #各グループのパス
        dnames_train = glob.glob('{}/*'.format(IMG_DIR_train))
        dnames_train = sorted(dnames_train, key=lambda s: int(re.search(r'\d+', s).group()))
        dnames_test = glob.glob('{}/*'.format(IMG_DIR_test))
        dnames_test = sorted(dnames_test, key=lambda s: int(re.search(r'\d+', s).group()))


        #画像ファイルパス一覧
        fnames_train = [glob.glob('{}/*.jpg'.format(d)) for d in dnames_train]
        fnames_train = list(chain.from_iterable(fnames_train))
        fnames_test = [glob.glob('{}/*.jpg'.format(d)) for d in dnames_test]
        fnames_test = list(chain.from_iterable(fnames_test))

        # それぞれにフォルダ名から一意なIDを付与し、画像を読み込んでデータセット作成
        labels_train = [os.path.basename(os.path.dirname(fn)) for fn in fnames_train]
        dnames_train = [os.path.basename(d) for d in dnames_train]
        labels_train = [dnames_train.index(l) for l in labels_train]
        d_train = LabeledImageDataset(list(zip(fnames_train, labels_train)))

        labels_test = [os.path.basename(os.path.dirname(fn)) for fn in fnames_test]
        dnames_test = [os.path.basename(d) for d in dnames_test]
        labels_test = [dnames_test.index(l) for l in labels_test]
        d_test = LabeledImageDataset(list(zip(fnames_test, labels_test)))

        #VGG用の前処理関数
        def transform(data):
            img, label = data
            img = L.model.vision.vgg.prepare(img, size=(h, w))
            img = img / 255. #正規化する．0〜1に落とし込む
            return img, label

        train = chainer.datasets.TransformDataset(d_train, transform)
        test = chainer.datasets.TransformDataset(d_test, transform)


        model = VGG(class_labels = G)



        #GPUの設定
        if gpu >= 0:
            print('GPUを用います')
            cuda.get_device(0).use()
            model.to_gpu(0)

        optimizer = optimizers.Adam() # 最適化関数
        optimizer.setup(model)
        # model.base.disable_update()


        #ラベルごとのトータル誤判別数
        miss_train_total = np.zeros(G)
        miss_test_total = np.zeros(G)

        #ラベルごとのトータル正解数
        correct_train_total = np.zeros(G)
        correct_test_total = np.zeros(G)

        #何を何と判別したか2次元配列
        miss_train_judge = np.zeros([G,G])#ここも
        miss_test_judge = np.zeros([G,G])

        cnt_train = 0#画像を保存するように使用
        cnt_test = 0#画像を保存するときの数
        cnt_train_correct_judge = 0
        cnt_test_correct_judge =0

        print(str(i)+'_trial')
        #エクセル書き込み
        sheet = book.add_sheet(str(i)+'_trial')
        sheet.write(1, 0, '教師用')
        sheet.write(2, 0, '評価用')
        sheet.write(6, 0, '教師用')
        sheet.write(6, G+3, '評価用')

        #葉齢ごとの判別率
        for l in range(0, G):
            sheet.write(7, l+1, 'stage'+str(l))

        for l in range(0, G):
            sheet.write(7, l+G+4, 'stage'+str(l))

        sheet.write(n_epoch+15, 0, '誤判別のやーつ(教師用)')
        sheet.write(n_epoch+15, 3, '誤判別先')
        for l in range(0,G):
            sheet.write(n_epoch+16, l+1, l)
            sheet.write(n_epoch+l+17, 0, l)

        sheet.write(n_epoch+15, G+3, '誤判別のやーつ(評価用)')
        for l in range(0, G):
            sheet.write(n_epoch+16, l+G+4, l)
            sheet.write(n_epoch+l+17, G+3, l)

        #trainer周りの設定
        train_iter = iterators.SerialIterator(train, batch_size)
        test_iter = iterators.SerialIterator(test, 1, repeat=False, shuffle=False)

        updater = training.StandardUpdater(train_iter, optimizer, device = gpu)
        trainer = training.Trainer(updater, (n_epoch, 'epoch'), out=current_result_dir)

        trainer.extend(extensions.dump_graph('main/loss'))
        #trainer.extend(extensions.snapshot(), trigger=(epoch, 'epoch'))
        trainer.extend(extensions.LogReport())
        trainer.extend(extensions.observe_lr())
        trainer.extend(extensions.PrintReport(
            ['epoch',
            'main/loss',
            'main/accuracy',
            'test/main/loss',
            'test/main/accuracy',
            'elapsed_time',
            'lr',], sheet = sheet, flag = 1))
        trainer.extend(extensions.ProgressBar(update_interval=1))
        trainer.extend(extensions.Evaluator(test_iter, model, device = gpu), name='test')
        
        trainer.extend(extensions.PlotReport(['main/loss'], 'epoch', file_name='main_loss.png', marker = ''))
        trainer.extend(extensions.PlotReport(['main/accuracy'], 'epoch', file_name='main_accuracy.png'))
        trainer.extend(extensions.PlotReport(['test/main/loss'], 'epoch', file_name='test_loss.png'))
        trainer.extend(extensions.PlotReport(['test/main/accuracy'], 'epoch', file_name='test_accuracy.png'))
        trainer.extend(extensions.PlotReport(['main/loss', 'test/main/loss'], x_key='epoch', file_name='loss.png', marker=""))
        trainer.extend(extensions.PlotReport(['main/accuracy', 'test/main/accuracy'], x_key='epoch', file_name='accuracy.png', marker=""))
        
        #trainer.extend(extensions.ExponentialShift('lr', lr_drop_ratio),trigger=(lr_drop_epoch, 'epoch'))
        
        trainer.run()

        paths = []

        for j in range(0, G):
            if j < 10:
                paths.append(np.asarray([os.path.join(os.path.join(test_dir, 'image2'), 'stage0'+str(j)+'\\'), j]))
            else:
                paths.append(np.asarray([os.path.join(os.path.join(test_dir, 'image2'), 'stage'+str(j)+'\\'), j]))

        (x_test, t_test), correct_test_label = image2Test(paths)



        #以下正確なaccuracyを求めるため
        test_loss = 0
        test_accuracy = 0
        miss_train_copy = miss_train_total.copy()
        miss_test_copy = miss_test_total.copy()
        correct_train_copy = correct_train_total.copy()
        correct_test_copy = correct_test_total.copy()

        for i in range(0, N_test, batch_size_predict):
            x = Variable(cuda.to_gpu(x_test[i:i+batch_size_predict]))#評価画像データ
            t = Variable(cuda.to_gpu(t_test[i:i+batch_size_predict]))#評価ラベル
            y = model.predict(x)
            model.zerograds()
            loss = F.softmax_cross_entropy(y, t)

            #こっから誤判別を特定するプログラム
            acc, data, data2, index, pred, corre, correct_index  = accuracy(y, t)
            acc.to_cpu()
            data.to_cpu()
            data2.to_cpu()
            index.to_cpu()
            pred.to_cpu()
            corre.to_cpu()
            correct_index.to_cpu()
            t.to_cpu()
            x.to_cpu()
            for i in range(len(data.data)):
                miss_test_total[data.data[i]] += 1
                miss_test_judge[data.data[i]][data2.data[i]] += 1

            for i in range(len(corre.data)):
                correct_test_total[corre.data[i]] += 1

            
            #こっからは誤判別した画像を出力するプログラム(最終更新時のみ)
            for inde in index.data:
                e = np.array(x.data[inde])
                b = np.array(e[0])#青色
                g = np.array(e[1])#緑色
                r = np.array(e[2])#赤色
                img = np.dstack((np.dstack((r,g)),b))

                #print(img)
                img = img * 255.0
                img += np.array([123.68, 116.779, 103.939], dtype=np.float32)
                img = img/255.0
                #print(img)
                plt.imsave(os.path.join(miss_judge_test_dir, str(cnt_trial)+'_trial\\'+str(n_epoch)+'\\'+str(cnt_test)+ '_epoch_' +str(n_epoch)+'_'+str(t.data[inde])+'to'+str(pred.data[inde])+'.jpg'), img)
                cnt_test += 1
            for correct_inde in correct_index.data:
                c_e = np.array(x.data[correct_inde])
                c_b = np.array(c_e[0])#青色
                c_g = np.array(c_e[1])#緑色
                c_r = np.array(c_e[2])#赤色
                c_img = np.dstack((np.dstack((c_r,c_g)),c_b))

                c_img = c_img * 255.0
                c_img += np.array([123.68, 116.779, 103.939], dtype=np.float32)
                c_img = c_img/255.0
                plt.imsave(os.path.join(correct_judge_test_dir, str(cnt_trial)+'_trial\\'+str(n_epoch)+'\\'+str(cnt_test_correct_judge)+'_epoch_'+str(n_epoch)+'_'+str(t.data[correct_inde])+'to'+str(pred.data[correct_inde])+'.jpg'), c_img)
                cnt_test_correct_judge += 1
            

            test_loss += loss.data*len(t.data)
            test_accuracy += acc.data*len(t.data)

            miss_test_each = miss_test_total - miss_test_copy
            correct_test_each = correct_test_total - correct_test_copy

        #ターミナル上に出力
        print("test acc:{:.04f}".format(test_accuracy/N_test))
        print('miss_test_total : ' + str(miss_test_total))
        print('miss_test_each :' + str(miss_test_each))
        print('correct_test_each :' + str(correct_test_each))
        print()

        # sheet.write(0, 1, n_epoch)#単純にエポック数(現時点)
        sheet.write(3, n_epoch, Decimal(test_accuracy/N_test).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))#評価用画像の判別率

        sheet.write((n_epoch + 7), 0, n_epoch)
        sheet.write((n_epoch + 7), G+3, n_epoch)
        for i in range(len(correct_test_each)):
            if correct_test_label[i] != 0:
                sheet.write((n_epoch + 7), (i + G+4), Decimal(correct_test_each[i]/correct_test_label[i]).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

        #最終epochの誤判定時における、何が何を予測したのかを表示
        for i in range(0, G):
            for j in range(0,G):
                if j==i:
                    sheet.write((n_epoch+17+i), (i+G+4), correct_test_each[i])
                else:
                    sheet.write((n_epoch+17+i), (j+G+4), miss_test_judge[i][j])

        #ここから先は前後のラベルを含んだ判別率を求める．
        num_zengo = np.zeros(G)#前後を含んだ正解数
        for i in range(0, G):
            if i == 0:
                num_zengo[i] = correct_test_each[i] + miss_test_judge[i][i+1]
            elif i == (G-1):
                num_zengo[i] = correct_test_each[i] + miss_test_judge[i][i-1]
            else:
                num_zengo[i] = correct_test_each[i] + miss_test_judge[i][i-1] + miss_test_judge[i][i+1]


        #num_zengoの出力とラベルごとの前後を含んだ判別率の出力
        for i in range(len(num_zengo)):
            if correct_test_label[i] != 0:
                sheet.write((n_epoch + 8), (i + G + 4), num_zengo[i])
                sheet.write((n_epoch + 9), (i + G + 4), Decimal(num_zengo[i]/correct_test_label[i]).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))

        #最後に前後を含んだ評価用画像の判別率を算出し，出力
        num_zengo_total = 0
        for i in range(0, G):
            num_zengo_total += num_zengo[i]
        sheet.write(4, n_epoch, Decimal(num_zengo_total/N_test).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))


        cnt_trial+=1
        end_time = time.clock()
        sheet.write(0,0,end_time-start_time)
        print(end_time - start_time)

        serializers.save_npz(os.path.join(current_result_dir, 'model.npz'), model)
        # shutil.rmtree(train_dir)#必要ならData augmentaion後の教師画像フォルダの削除

        book.save(os.path.join(result_dir, 'result'+str(cnt_trial-1)+'_trial.xls'))

        change_excel(os.path.join(result_dir, 'result'+str(cnt_trial-1)+'_trial.xls'))
        os.remove(os.path.join(result_dir, 'result'+str(cnt_trial-1)+'_trial.xls'))

    excel_names = glob.glob(os.path.join(result_dir, "*.xlsx"))
    ordered_excel = natsorted(excel_names)

    outfile = os.path.join(result_dir, 'result.xlsx')
    wb2 = xl.Workbook()
    wb2.save(outfile)
    print(str(ordered_excel))
    for i, inputfile in enumerate(ordered_excel):
        print(str(i))
        sheettitle = str(i+1) + '_trial'
        print(sheettitle)
        print(str(inputfile))

        wb1 = xl.load_workbook(filename = inputfile)
        ws1 = wb1.worksheets[0]

        wb = xl.load_workbook(filename = outfile)
        if i > 0:
            wb.create_sheet()
        ws = wb.worksheets[i]
        ws.title = sheettitle

        for row in ws1:
            for cell in row:
                ws[cell.coordinate].value = cell.value

        wb.save(outfile)